import openpyxl as pyxl
from openpyxl.styles import NamedStyle

RESULT_SHEET_NAME = "Relatório"

if __name__ == "__main__":
    wbook = pyxl.Workbook()

    try:
        wbook = pyxl.load_workbook("./Exemplo.xlsm")
    except FileNotFoundError as id:
        print("Planilha não encontrada!", id)
    else:
        wsheet = wbook[wbook.sheetnames[0]]
        print("Planilha Carregada!")

    sheets = []
    sheet_names = []
    for name in wbook.sheetnames:
        sheets.append(wbook[name])
        sheet_names.append(name)
    print("Abas encontradas:", wbook.sheetnames, "\n")

    wbook.create_sheet(RESULT_SHEET_NAME, 0)
    result_sheet = wbook[RESULT_SHEET_NAME]

    for row in sheets[0].iter_rows(1, 1):
        header = [cell.value for cell in row]
    header.extend(["Frota"])
    result_sheet.append(header)

#      for n, sheet in enumerate(sheets[0:-1]):
    for n, sheet in enumerate(sheets):
        for row_n, row in enumerate(sheet.iter_rows()):
            if row_n > 0:
                row_values = [cell.value for cell in row]
                row_values.extend([sheet_names[n]])
                result_sheet.append(row_values)

    custom = NamedStyle(name='custom_datetime', number_format='dd/mmm')
    for col in result_sheet.iter_cols(1, 1, 2):
        for cell in col:
            cell.style = custom

    wbook.save("./Resultado.xlsm")
